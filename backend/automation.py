import schedule
import time
import threading
from datetime import datetime
from database import get_all_entries
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import smtplib
from email.message import EmailMessage
import mimetypes

# Настройки email
EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',  # Или smtp.yandex.ru для Яндекс
    'smtp_port': 465,
    'sender_email': 'mr.replicate@gmail.com',  # Твой email
    'sender_password': 'uzopnnkinofqxzsp',  # Пароль приложения
    'recipient_email': 'Irvees78@gmail.com'  # Куда отправлять
}

def generate_auto_report():
    """Функция, которая создаёт отчёт автоматически"""
    print(f"\n🤖 [АВТОМАТИЗАЦИЯ] Запуск генерации отчёта: {datetime.now()}")
    
    # Получаем все записи
    all_entries = get_all_entries()
    
    # Фильтруем только за сегодня
    today = datetime.now().strftime("%Y-%m-%d")
    entries = [e for e in all_entries if e[3] and e[3].startswith(today)]
    
    # Если записей за сегодня нет — не создаём отчёт
    if not entries:
        print(f"⚠ [АВТОМАТИЗАЦИЯ] Записей за {today} нет, отчёт не создан")
        return
    
    # Считаем общую статистику
    total_seconds = sum([e[5] for e in entries if e[5]])
    total_hours = round(total_seconds / 3600, 2)
    active_projects = len(set([e[1] for e in entries]))
    
    # Создаём папку, если нет
    os.makedirs('reports/auto', exist_ok=True)
    
    # Создаём Excel книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Авто-отчёт"
    
    # Заголовок
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = f"🤖 АВТОМАТИЧЕСКИЙ ОТЧЁТ ЗА {today}"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Статистика
    ws.append(['Дата', today, ''])
    ws.append(['Всего проектов', active_projects, ''])
    ws.append(['Всего часов', total_hours, ''])
    ws.append([])
    
    # Заголовки таблицы
    headers = ['ID', 'Проект', 'Описание', 'Начало', 'Конец', 'Секунды']
    ws.append(headers)
    
    # Стиль шапки
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    
    # Данные
    for e in entries:
        if e[5]:
            start = e[3].replace('T', ' ') if e[3] else ''
            end = e[4].replace('T', ' ') if e[4] else ''
            ws.append([e[0], e[1], e[2] or '', start, end, e[5]])
    
    # Имя файла
    filename = f'reports/auto/DAILY_REPORT_{today}.xlsx'
    
    # Сохраняем
    wb.save(filename)

    # Отправляем на email
    send_email_report(filename)
    
    print(f"✅ [АВТОМАТИЗАЦИЯ] Excel-отчёт сохранён: {filename}")
    print(f"📊 Итого за {today}: {total_hours} ч. ({len(entries)} записей)")

def start_scheduler():
    """Запускает планировщик в фоновом режиме"""
    
    # Настраиваем расписание (время сервера)
    schedule.every().day.at("18:00").do(generate_auto_report)
    schedule.every().day.at("09:00").do(generate_auto_report)
    schedule.every().day.at("02:05").do(generate_auto_report)
    
    print("⏰ Планировщик запущен. Ожидание задач...")
    
    while True:
        schedule.run_pending()
        time.sleep(1)

def send_email_report(filename):
    """Отправляет отчёт на email"""
    
    print(f"📧 [EMAIL] Отправка отчёта на {EMAIL_CONFIG['recipient_email']}...")
    
    try:
        # Создаём сообщение
        msg = EmailMessage()
        msg['Subject'] = f'📊 Отчёт за {datetime.now().strftime("%d.%m.%Y")}'
        msg['From'] = EMAIL_CONFIG['sender_email']
        msg['To'] = EMAIL_CONFIG['recipient_email']
        msg.set_content(f'''
Здравствуйте!

Автоматический отчёт по учёту рабочего времени за {datetime.now().strftime("%d.%m.%Y")}.

Всего часов: {sum([e[5] for e in get_all_entries() if e[5] and e[3].startswith(datetime.now().strftime("%Y-%m-%d"))]) / 3600:.2f}

Файл с деталями во вложении.

---
Система учёта времени
        ''')
        
        # Прикрепляем файл
        with open(filename, 'rb') as f:
            file_data = f.read()
            file_name = os.path.basename(filename)
        
        msg.add_attachment(
            file_data,
            maintype='application',
            subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=file_name
        )
        
        # Отправляем (используем SSL)
        with smtplib.SMTP_SSL(EMAIL_CONFIG['smtp_server'], 465) as server:
            server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
            server.send_message(msg)
        
        print(f"✅ [EMAIL] Отчёт успешно отправлен!")
        return True
        
    except Exception as e:
        print(f"❌ [EMAIL] Ошибка отправки: {e}")
        return False
    
def run_scheduler_in_background():
    """Запускает планировщик отдельно, чтобы не блокировать сайт"""
    thread = threading.Thread(target=start_scheduler, daemon=True)
    thread.start()
    print("🔄 Фоновый режим активирован")
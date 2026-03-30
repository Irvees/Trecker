from flask import Flask, render_template, request, jsonify, send_file
from database import add_entry, close_entry, get_all_entries
from automation import run_scheduler_in_background
from datetime import datetime
import csv
import os
import sqlite3

# Определяем пути к папкам
app_folder = os.path.dirname(os.path.abspath(__file__))
parent_folder = os.path.dirname(app_folder)

# Создаём Flask приложение с правильными путями
app = Flask(__name__,
            template_folder=os.path.join(parent_folder, 'templates'),
            static_folder=os.path.join(parent_folder, 'static'))

# Главная страница
@app.route('/')
def index():
    """Показывает главную страницу с таймером"""
    return render_template('index.html')

# Запуск таймера
@app.route('/api/start', methods=['POST'])
def start_timer():
    """Записывает начало работы"""
    data = request.json
    project = data.get('project', 'Без названия')
    description = data.get('description', '')
    start_time = datetime.now().isoformat()
    
    add_entry(project, description, start_time)
    
    return jsonify({
        'status': 'success',
        'message': 'Таймер запущен!',
        'start_time': start_time
    })

# Остановка таймера
@app.route('/api/stop', methods=['POST'])
def stop_timer():
    """Записывает конец работы и считает длительность"""
    conn = sqlite3.connect('timetracker.db')
    cursor = conn.cursor()
    
    # Находим последнюю незакрытую запись
    cursor.execute('''
        SELECT id, start_time FROM entries 
        WHERE end_time IS NULL 
        ORDER BY id DESC LIMIT 1
    ''')
    entry = cursor.fetchone()
    conn.close()
    
    if entry:
        entry_id, start_time = entry
        end_time = datetime.now()
        start_dt = datetime.fromisoformat(start_time)
        duration = int((end_time - start_dt).total_seconds())
        
        close_entry(entry_id, end_time.isoformat(), duration)
        
        return jsonify({
            'status': 'success',
            'message': 'Таймер остановлен!',
            'duration': f'{duration} сек. ({duration // 60} мин.)'
        })
    else:
        return jsonify({
            'status': 'error',
            'message': 'Нет активных таймеров!'
        })

# Получение всех записей
@app.route('/api/entries', methods=['GET'])
def get_entries():
    """Возвращает все записи для отображения"""
    entries = get_all_entries()
    result = []
    for e in entries:
        result.append({
            'id': e[0],
            'project': e[1],
            'description': e[2],
            'start_time': e[3],
            'end_time': e[4],
            'duration': e[5]
        })
    return jsonify(result)

# Экспорт в CSV (ручной)
@app.route('/api/export', methods=['GET'])
def export_excel():
    """Создаёт красивый Excel-отчёт"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    entries = get_all_entries()
    
    # Создаём Excel книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по времени"
    
    # Заголовок
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "📊 ОТЧЁТ ПО УЧЁТУ ВРЕМЕНИ"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Шапка таблицы
    headers = ['ID', 'Проект', 'Описание', 'Начало', 'Конец', 'Длительность (сек)']
    ws.append(headers)
    
    # Стиль шапки
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for e in entries:
        if e[5]:  # Только завершённые записи
            # Форматируем время
            start = e[3].replace('T', ' ') if e[3] else ''
            end = e[4].replace('T', ' ') if e[4] else ''
            ws.append([e[0], e[1], e[2] or '', start, end, e[5]])
    
    # Применяем границы ко всем ячейкам
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Авто-ширина колонок
    column_widths = [10, 25, 30, 22, 22, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Итоговая статистика
    total_seconds = sum([e[5] for e in entries if e[5]])
    total_hours = round(total_seconds / 3600, 2)
    total_projects = len(set([e[1] for e in entries]))
    
    ws.append([])
    ws.append(['📈 ИТОГО:', '', '', '', '', ''])
    ws.append(['Всего записей:', len(entries), '', '', '', ''])
    ws.append(['Всего проектов:', total_projects, '', '', '', ''])
    ws.append(['Всего часов:', total_hours, '', '', '', ''])
    
    # Стиль итогов
    for row in range(ws.max_row - 3, ws.max_row + 1):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Сохраняем
    os.makedirs('reports', exist_ok=True)
    filename = f'reports/report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    
    return jsonify({
        'status': 'success',
        'message': f'Excel-отчёт сохранён: {filename}'
    })

# Запуск приложения
if __name__ == '__main__':
    print("🚀 Запуск сервера...")
    
    # 👇 Запускаем автоматизацию в фоновом режиме
    run_scheduler_in_background()
    # 👆 Конец запуска автоматизации
    
    print("📍 Откройте в браузере: http://localhost:5000")
    print("🤖 Автоматизация активна (отчёты в 09:00 и 18:00)")
    app.run(debug=True, use_reloader=False)